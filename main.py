from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from typing import Optional, Dict, List, Any
import base64
from io import BytesIO
import json

import httpx
from openpyxl import load_workbook

app = FastAPI(
    title="Excel to JSON Service",
    description="Convert Excel files (via URL or base64) to JSON for Coze plugins.",
    version="1.0.0",
)


class ConvertRequest(BaseModel):
    file_url: Optional[str] = None
    file_base64: Optional[str] = None
    sheet_name: Optional[str] = None
    header_row: bool = True


class ConvertResponse(BaseModel):
    sheets: str


class ConvertQAResponse(BaseModel):
    items: str


async def _download_file(url: str) -> bytes:
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            return resp.content
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to download file from URL: {e}")


def _load_excel(content: bytes):
    try:
        return load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}")


def _sheet_to_rows(ws, header_row: bool) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    if header_row:
        headers = [str(c) if c is not None else f"col_{idx + 1}" for idx, c in enumerate(rows[0])]
        data_rows = rows[1:]
    else:
        max_len = max(len(r) for r in rows)
        headers = [f"col_{idx + 1}" for idx in range(max_len)]
        data_rows = rows

    result: List[Dict[str, Any]] = []
    for row in data_rows:
        item: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else None
            item[header] = value
        result.append(item)

    return result


@app.get("/health")
async def health_check():
    return {"status": "ok"}


async def _build_result(
    file_url: Optional[str],
    file_base64: Optional[str],
    sheet_name: Optional[str],
    header_row: bool,
) -> Dict[str, List[Dict[str, Any]]]:
    if not file_url and not file_base64:
        raise HTTPException(status_code=400, detail="Either file_url or file_base64 must be provided.")

    if file_url:
        content = await _download_file(file_url)
    else:
        try:
            content = base64.b64decode(file_base64 or "")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid base64 content: {e}")

    wb = _load_excel(content)

    result: Dict[str, List[Dict[str, Any]]] = {}

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found in workbook.")
        ws = wb[sheet_name]
        result[ws.title] = _sheet_to_rows(ws, header_row)
    else:
        for ws in wb.worksheets:
            result[ws.title] = _sheet_to_rows(ws, header_row)

    return result


def _build_qa_items(result: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    qa_items: List[Dict[str, Any]] = []

    for sheet_name, rows in result.items():
        if not rows:
            continue

        # 第一行是题目行
        question_row = rows[0]

        # 后面的每一行是学员答案
        for answer_row in rows[1:]:
            name = answer_row.get("姓名")
            score = answer_row.get("客观题得分")

            for field, question in question_row.items():
                if field in ("姓名", "客观题得分"):
                    continue

                answer = answer_row.get(field)

                qa_items.append(
                    {
                        "sheet": sheet_name,
                        "name": name,
                        "field": field,
                        "question": question,
                        "answer": answer,
                        "score": score,
                    }
                )

    return qa_items


async def _convert(
    file_url: Optional[str],
    file_base64: Optional[str],
    sheet_name: Optional[str],
    header_row: bool,
) -> ConvertResponse:
    result = await _build_result(file_url, file_base64, sheet_name, header_row)

    # Serialize sheets dict to JSON string so it matches the OpenAPI schema (string)
    return ConvertResponse(sheets=json.dumps(result, ensure_ascii=False))


@app.post("/convert", response_model=ConvertResponse)
async def convert_excel(req: ConvertRequest):
    return await _convert(req.file_url, req.file_base64, req.sheet_name, req.header_row)


@app.get("/convert", response_model=ConvertResponse)
async def convert_excel_query(
    file_url: Optional[str] = None,
    file_base64: Optional[str] = None,
    sheet_name: Optional[str] = None,
    header_row: bool = True,
):
    return await _convert(file_url, file_base64, sheet_name, header_row)


@app.get("/convert_qa", response_model=ConvertQAResponse)
async def convert_excel_qa_query(
    file_url: Optional[str] = None,
    file_base64: Optional[str] = None,
    sheet_name: Optional[str] = None,
    header_row: bool = True,
):
    result = await _build_result(file_url, file_base64, sheet_name, header_row)
    qa_items = _build_qa_items(result)
    return ConvertQAResponse(items=json.dumps(qa_items, ensure_ascii=False))


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8001)
